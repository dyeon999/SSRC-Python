from openpyxl import Workbook

wb = Workbook()
ws = wb.active #워크시트
ws. title = "New Sheet"

# for row in ws['A1:D4']:
#     for cell in row:
#         cell.value = "GoodMorning!"

for i in range(10):
    row_cell = ws.cell(row=(i+1), column=2)
    row_cell.value = str(i+1) + "번째 데이터 저장"

wb.save("Day03/test.xlsx")